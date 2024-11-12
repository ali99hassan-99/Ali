from flask import Flask, request, render_template
import openpyxl

app = Flask(__name__)

# دالة لتحميل ملف Excel أو إنشاء ملف جديد
def load_workbook():
    try:
        # إذا كان الملف موجودًا، نفتح الملف
        return openpyxl.load_workbook('orders.xlsx')
    except FileNotFoundError:
        # إذا لم يكن الملف موجودًا، ننشئ ملفًا جديدًا
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['اسم العميل', 'اسم المنتج', 'الكمية', 'السعر', 'المجموع'])  # رؤوس الأعمدة
        wb.save('orders.xlsx')
        return wb

# عرض صفحة الطلبات
@app.route('/')
def index():
    return render_template('index.html')

# معالجة البيانات بعد تقديم الطلب
@app.route('/submit_order', methods=['POST'])
def submit_order():
    # جمع البيانات من النموذج
    customer_name = request.form['customer_name']
    product_name = request.form['product_name']
    quantity = int(request.form['quantity'])
    price = float(request.form['price'])
    total = quantity * price

    # إضافة البيانات إلى ملف Excel
    wb = load_workbook()
    sheet = wb.active
    sheet.append([customer_name, product_name, quantity, price, total])  # إضافة البيانات الجديدة
    wb.save('orders.xlsx')  # حفظ الملف

    # توجيه المستخدم إلى صفحة تأكيد الطلب
    return render_template('order_confirmation.html', customer_name=customer_name, total=total)

# تشغيل السيرفر
if __name__ == '__main__':
    app.run(debug=True)
